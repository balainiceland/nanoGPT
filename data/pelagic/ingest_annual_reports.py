#!/usr/bin/env python3
"""Ingest annual report PDFs into PelagicGPT training corpus.

Reads the annual_reports_database.xlsx, downloads each PDF, extracts text,
chunks it with mode tags, and appends to the training data.

Usage:
    python ingest_annual_reports.py                # Preview (no tokenize)
    python ingest_annual_reports.py --tokenize     # Rebuild train.bin / val.bin
"""

import argparse
import os
import sys
import re
from pathlib import Path

import openpyxl

XLSX_PATH = Path('/Users/Bala_1/Downloads/annual_reports_database.xlsx')
PDF_DIR = Path(__file__).parent / 'annual_reports'
CORPUS_FILE = Path(__file__).parent / 'annual_reports_corpus.txt'
EOT = '<|endoftext|>'

# Subsector → mode tag mapping
SUBSECTOR_TAGS = {
    'Aquaculture': '[MARKET COMMENTARY]',
    'Ocean Tech': '[MARKET COMMENTARY]',
    'Shipping': '[MARKET COMMENTARY]',
    'Water Tech': '[MARKET COMMENTARY]',
    'Marine Biotech': '[MARKET COMMENTARY]',
    'Marine Infrastructure': '[MARKET COMMENTARY]',
    'Offshore Wind': '[MARKET COMMENTARY]',
}


def extract_reports_from_xlsx() -> list[dict]:
    """Parse the Excel file and return list of report metadata with URLs."""
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb['Annual Reports Database']

    reports = []
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
        company = row[2].value   # Column C
        if not company:
            continue

        subsector = row[1].value or 'Other'
        ticker = row[3].value or ''
        weight = row[5].value or ''
        report_title = row[7].value or ''
        pdf_url = row[9].hyperlink.target if row[9].hyperlink else None
        ir_url = row[10].hyperlink.target if row[10].hyperlink else None

        if not pdf_url:
            print(f"  WARNING: No PDF URL for {company}, skipping")
            continue

        reports.append({
            'company': company,
            'ticker': ticker,
            'subsector': subsector,
            'weight': weight,
            'report_title': report_title,
            'pdf_url': pdf_url,
            'ir_url': ir_url,
        })

    return reports


def download_pdf(url: str, dest: Path) -> bool:
    """Download PDF if not already cached."""
    if dest.exists() and dest.stat().st_size > 10000:
        return True

    import urllib.request
    try:
        req = urllib.request.Request(url, headers={
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)',
        })
        with urllib.request.urlopen(req, timeout=60) as resp:
            dest.write_bytes(resp.read())
        return True
    except Exception as e:
        print(f"  FAILED to download: {e}")
        return False


def extract_text_from_pdf(pdf_path: Path) -> str:
    """Extract text from PDF using PyMuPDF (fitz)."""
    try:
        import fitz  # PyMuPDF
    except ImportError:
        print("ERROR: pip install PyMuPDF")
        sys.exit(1)

    doc = fitz.open(str(pdf_path))
    pages = []
    for page in doc:
        text = page.get_text()
        if text.strip():
            pages.append(text.strip())
    doc.close()
    return '\n\n'.join(pages)


def clean_text(text: str) -> str:
    """Clean extracted PDF text for training."""
    # Remove excessive whitespace
    text = re.sub(r'\n{3,}', '\n\n', text)
    # Remove page numbers / headers that are just numbers
    text = re.sub(r'^\d{1,3}\s*$', '', text, flags=re.MULTILINE)
    # Remove very short lines (likely headers/footers)
    lines = text.split('\n')
    cleaned = []
    for line in lines:
        stripped = line.strip()
        # Keep lines with substance
        if len(stripped) > 20 or stripped == '':
            cleaned.append(line)
    return '\n'.join(cleaned)


def chunk_text(text: str, company: str, ticker: str, subsector: str,
               report_title: str, chunk_size: int = 2000) -> list[str]:
    """Split text into training chunks with mode tags."""
    tag = SUBSECTOR_TAGS.get(subsector, '[MARKET COMMENTARY]')
    paragraphs = text.split('\n\n')
    chunks = []
    current = []
    current_len = 0

    for para in paragraphs:
        para = para.strip()
        if not para or len(para) < 30:
            continue

        if current_len + len(para) > chunk_size and current:
            header = f"{tag}\n{company} ({ticker}) — {report_title}\n"
            chunk_text = header + '\n\n'.join(current)
            chunks.append(f"{EOT}\n{chunk_text}")
            current = []
            current_len = 0

        current.append(para)
        current_len += len(para)

    if current:
        header = f"{tag}\n{company} ({ticker}) — {report_title}\n"
        chunk_text = header + '\n\n'.join(current)
        chunks.append(f"{EOT}\n{chunk_text}")

    return chunks


def main():
    parser = argparse.ArgumentParser(description='Ingest annual reports for PelagicGPT')
    parser.add_argument('--tokenize', action='store_true', help='Rebuild train.bin/val.bin')
    args = parser.parse_args()

    print("=== Annual Reports Ingestion for PelagicGPT ===\n")

    # 1. Parse Excel
    reports = extract_reports_from_xlsx()
    print(f"Found {len(reports)} reports in database\n")

    # 2. Download & extract
    PDF_DIR.mkdir(exist_ok=True)
    all_chunks = []

    for r in reports:
        safe_name = re.sub(r'[^\w\-.]', '_', f"{r['ticker']}_{r['report_title'][:40]}")
        pdf_path = PDF_DIR / f"{safe_name}.pdf"

        print(f"  {r['ticker']:12} {r['company'][:25]:25}", end=' ')

        if download_pdf(r['pdf_url'], pdf_path):
            size_mb = pdf_path.stat().st_size / (1024 * 1024)
            text = extract_text_from_pdf(pdf_path)
            cleaned = clean_text(text)
            chunks = chunk_text(
                cleaned,
                company=r['company'],
                ticker=r['ticker'],
                subsector=r['subsector'],
                report_title=r['report_title'],
            )
            all_chunks.extend(chunks)
            print(f"  {size_mb:5.1f} MB  {len(text):>8,} chars  {len(chunks):>4} chunks")
        else:
            print(f"  SKIPPED")

    # 3. Save corpus
    corpus = '\n'.join(all_chunks) + f'\n{EOT}\n'
    CORPUS_FILE.write_text(corpus, encoding='utf-8')
    print(f"\nTotal: {len(all_chunks)} chunks, {len(corpus):,} characters")
    print(f"Saved to {CORPUS_FILE.name}")

    # 4. Optionally tokenize and merge with existing training data
    if args.tokenize:
        import numpy as np
        import tiktoken

        print("\nTokenizing...")
        enc = tiktoken.get_encoding("gpt2")
        new_tokens = enc.encode(corpus, allowed_special={EOT})
        print(f"Annual report tokens: {len(new_tokens):,}")

        output_dir = Path(__file__).parent
        train_path = output_dir / 'train.bin'
        val_path = output_dir / 'val.bin'

        if train_path.exists():
            existing_train = np.fromfile(train_path, dtype=np.uint16)
            existing_val = np.fromfile(val_path, dtype=np.uint16)
            all_existing = np.concatenate([existing_train, existing_val])
            combined = np.concatenate([all_existing, np.array(new_tokens, dtype=np.uint16)])
            print(f"Existing tokens: {len(all_existing):,}")
            print(f"Combined tokens: {len(combined):,}")
        else:
            combined = np.array(new_tokens, dtype=np.uint16)

        n = len(combined)
        split_idx = int(n * 0.9)
        np.array(combined[:split_idx]).tofile(train_path)
        np.array(combined[split_idx:]).tofile(val_path)

        print(f"\ntrain.bin: {split_idx:,} tokens")
        print(f"val.bin:   {n - split_idx:,} tokens")
        print("Ready for training!")


if __name__ == '__main__':
    main()
